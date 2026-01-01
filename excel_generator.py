import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from flask_mysqldb import MySQL
import MySQLdb.cursors
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self, reports_dir='reports'):
        self.reports_dir = reports_dir
        os.makedirs(reports_dir, exist_ok=True)
    
    def generate_attendance_report(self, classroom_id, mysql_connection=None):
        """
        Generate Excel attendance report for a classroom
        Returns: filepath or None
        """
        try:
            if not mysql_connection:
                logger.error("MySQL connection not provided")
                return None
            
            cursor = mysql_connection.cursor(MySQLdb.cursors.DictCursor)
            
            # Get classroom details
            cursor.execute('''
                SELECT c.class_name, c.class_code, c.semester,
                       t.full_name as teacher_name
                FROM classrooms c
                JOIN teachers t ON c.teacher_id = t.id
                WHERE c.id = %s
            ''', (classroom_id,))
            classroom = cursor.fetchone()
            
            if not classroom:
                logger.error(f"Classroom {classroom_id} not found")
                return None
            
            # Get all students in classroom
            cursor.execute('''
                SELECT s.id, s.roll_number, s.full_name, s.email,
                       s.department, s.year
                FROM students s
                JOIN classroom_students cs ON s.id = cs.student_id
                WHERE cs.classroom_id = %s
                ORDER BY s.roll_number
            ''', (classroom_id,))
            students = cursor.fetchall()
            
            if not students:
                logger.warning(f"No students found in classroom {classroom_id}")
                return None
            
            # Get attendance dates for this classroom
            cursor.execute('''
                SELECT DISTINCT date
                FROM attendance
                WHERE classroom_id = %s
                ORDER BY date
            ''', (classroom_id,))
            dates = [row['date'] for row in cursor.fetchall()]
            
            # Create DataFrame
            data = []
            for student in students:
                student_data = {
                    'Roll No': student['roll_number'],
                    'Name': student['full_name'],
                    'Department': student['department'],
                    'Year': student['year'],
                    'Email': student['email']
                }
                
                # Get attendance for each date
                for date in dates:
                    cursor.execute('''
                        SELECT status FROM attendance
                        WHERE student_id = %s AND classroom_id = %s AND date = %s
                    ''', (student['id'], classroom_id, date))
                    
                    attendance = cursor.fetchone()
                    status = attendance['status'] if attendance else 'A'
                    student_data[date.strftime('%Y-%m-%d')] = status
                
                # Calculate totals
                cursor.execute('''
                    SELECT 
                        COUNT(*) as total_days,
                        SUM(CASE WHEN status = 'P' THEN 1 ELSE 0 END) as present_days,
                        SUM(CASE WHEN status = 'A' THEN 1 ELSE 0 END) as absent_days,
                        SUM(CASE WHEN status = 'L' THEN 1 ELSE 0 END) as late_days
                    FROM attendance
                    WHERE student_id = %s AND classroom_id = %s
                ''', (student['id'], classroom_id))
                
                totals = cursor.fetchone()
                total = totals['total_days'] or 0
                present = totals['present_days'] or 0
                
                student_data['Total Days'] = total
                student_data['Present'] = present
                student_data['Absent'] = totals['absent_days'] or 0
                student_data['Late'] = totals['late_days'] or 0
                student_data['Attendance %'] = round((present / total * 100), 2) if total > 0 else 0
                
                data.append(student_data)
            
            cursor.close()
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"attendance_{classroom['class_code']}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Attendance sheet
                df.to_excel(writer, sheet_name='Attendance', index=False)
                
                # Summary sheet
                summary_data = {
                    'Metric': ['Total Students', 'Total Days', 'Average Attendance %'],
                    'Value': [
                        len(students),
                        len(dates),
                        round(df['Attendance %'].mean(), 2) if not df.empty else 0
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format the workbook
                workbook = writer.book
                attendance_sheet = writer.sheets['Attendance']
                
                # Set column widths
                for column in attendance_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    attendance_sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add conditional formatting for attendance
                from openpyxl.formatting.rule import FormulaRule
                from openpyxl.styles import PatternFill, Font
                
                # Green for Present
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                present_rule = FormulaRule(formula=['$1="P"'], fill=green_fill)
                attendance_sheet.conditional_formatting.add('F2:Z1000', present_rule)
                
                # Red for Absent
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                absent_rule = FormulaRule(formula=['$1="A"'], fill=red_fill)
                attendance_sheet.conditional_formatting.add('F2:Z1000', absent_rule)
                
                # Yellow for Late
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                late_rule = FormulaRule(formula=['$1="L"'], fill=yellow_fill)
                attendance_sheet.conditional_formatting.add('F2:Z1000', late_rule)
            
            logger.info(f"Report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return None
    
    def generate_detailed_report(self, classroom_id, start_date, end_date, mysql_connection):
        """Generate detailed report with date range"""
        try:
            cursor = mysql_connection.cursor(MySQLdb.cursors.DictCursor)
            
            # Get detailed attendance
            cursor.execute('''
                SELECT 
                    s.roll_number,
                    s.full_name,
                    a.date,
                    a.status,
                    a.marked_at,
                    a.verification_method,
                    a.confidence_score
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.classroom_id = %s 
                    AND a.date BETWEEN %s AND %s
                ORDER BY a.date, s.roll_number
            ''', (classroom_id, start_date, end_date))
            
            records = cursor.fetchall()
            cursor.close()
            
            if not records:
                return None
            
            # Create DataFrame
            df = pd.DataFrame(records)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"detailed_report_{start_date}_{end_date}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Save to Excel
            df.to_excel(filepath, index=False)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating detailed report: {e}")
            return None